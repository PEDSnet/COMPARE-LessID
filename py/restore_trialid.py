#!/usr/bin/env python3
"""
restore_trialid.py  [--dry-run]  [site ...]

Restores original participant/trial IDs in de-identified XLSX output files.

The previous pipeline incorrectly remapped columns whose XLSX headers resolve
to the CDM column "trialid" (c_partid, e_partid, c_trialid). This script
uses the existing mapping.csv (which records original_value → new_id) to
build the reverse map and patch XLSX files in-place.

Columns affected (via XLSX_COLUMN_MAP):
  "c_partid"  → trialid
  "e_partid"  → trialid
  "c_trialid" → trialid

Usage:
  python3 py/restore_trialid.py                    # all sites
  python3 py/restore_trialid.py C7LC_compare_deq_q01 C7CNH_compare_deq_q01
  python3 py/restore_trialid.py --dry-run           # preview only, no writes
"""

import csv
import glob
import os
import sys
import warnings
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rules import patch_openpyxl
patch_openpyxl()

# ── Paths ────────────────────────────────────────────────────────────────────
OUT_BASE    = os.environ.get("OUT_BASE",    "/data/sas_queries/seyediana/lessid_drnoc")
LOOKUP_BASE = os.environ.get("LOOKUP_BASE", "/data/sas_queries/seyediana/lessid_lookup")

# XLSX column headers (lowercased) that carry participant/datamart IDs and must be restored
TRIALID_XLSX_HEADERS = {"participant id", "datamart id", "c_partid", "e_partid", "c_trialid"}

# ── Args ─────────────────────────────────────────────────────────────────────
args = sys.argv[1:]
dry_run = "--dry-run" in args
sites_arg = [a for a in args if not a.startswith("--")]

if sites_arg:
    sites = sites_arg
else:
    if not os.path.isdir(OUT_BASE):
        print(f"ERROR: OUT_BASE not found: {OUT_BASE}")
        sys.exit(1)
    sites = sorted(os.listdir(OUT_BASE))

if dry_run:
    print("DRY RUN — no files will be written\n")

total_restored = 0
total_missing  = 0

for site in sites:
    out_dir     = os.path.join(OUT_BASE,    site)
    mapping_csv = os.path.join(LOOKUP_BASE, site, "mapping.csv")

    if not os.path.isdir(out_dir):
        print(f"[{site}] SKIP — output dir not found: {out_dir}")
        continue
    if not os.path.isfile(mapping_csv):
        print(f"[{site}] SKIP — mapping.csv not found: {mapping_csv}")
        continue

    # Build reverse map: new_id → original_value for participant ID columns.
    # The mapping CSV keys these as "participant id" and "datamart id" (with spaces)
    # matching the raw XLSX column header lowercased.
    PARTICIPANT_COL_KEYS = {"participant id", "datamart id", "c_partid", "e_partid", "c_trialid"}
    reverse_map: dict[str, str] = {}
    with open(mapping_csv, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if (row.get("column") or "").strip().lower() in PARTICIPANT_COL_KEYS:
                new_id = (row.get("new_id") or "").strip()
                orig   = (row.get("original_value") or "").strip()
                if new_id and orig:
                    reverse_map[new_id] = orig

    if not reverse_map:
        print(f"[{site}] no c_partid/e_partid entries in mapping.csv — nothing to restore")
        continue

    print(f"[{site}] {len(reverse_map)} participant ID reverse mappings loaded")

    xlsx_files = sorted(glob.glob(os.path.join(out_dir, "*.xlsx")))
    if not xlsx_files:
        print(f"[{site}] no XLSX files found")
        continue

    site_restored = 0
    site_missing  = 0

    for xlsx_path in xlsx_files:
        fname = os.path.basename(xlsx_path)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(xlsx_path)
        except Exception as e:
            print(f"  [{fname}] ERROR opening: {e}")
            continue

        file_restored = 0
        file_missing  = 0

        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue

            headers = [
                (str(cell.value).strip().lower() if cell.value is not None else "")
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            trialid_cols = [
                i for i, h in enumerate(headers)
                if h in TRIALID_XLSX_HEADERS
            ]
            if not trialid_cols:
                continue

            for row in ws.iter_rows(min_row=2):
                for i in trialid_cols:
                    v = row[i].value
                    if v is None:
                        continue
                    s = str(v).strip()
                    if not s:
                        continue
                    orig = reverse_map.get(s)
                    if orig is None:
                        # Not a remapped value — either already original or unknown
                        file_missing += 1
                        continue
                    if not dry_run:
                        row[i].value = orig
                    file_restored += 1

        if file_restored or file_missing:
            status = "would restore" if dry_run else "restored"
            print(f"  [{fname}] {status} {file_restored} values"
                  + (f" ({file_missing} already-original or unrecognised)" if file_missing else ""))
        else:
            print(f"  [{fname}] no trialid columns found")

        if not dry_run and file_restored:
            wb.save(xlsx_path)

        site_restored += file_restored
        site_missing  += file_missing

    print(f"[{site}] total: {site_restored} restored, {site_missing} skipped\n")
    total_restored += site_restored
    total_missing  += site_missing

print(f"Done. Grand total: {total_restored} values restored across all sites.")
