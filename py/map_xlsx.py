#!/usr/bin/env python3
import csv
import sys
import os
import warnings
import datetime
import openpyxl
import openpyxl.descriptors.base as _openpyxl_base

if len(sys.argv) != 4:
    print("Usage: map_xlsx.py <mapping.csv> <input.xlsx> <output.xlsx>")
    sys.exit(1)

MAPPING_CSV = sys.argv[1]
INPUT_PATH = sys.argv[2]
OUTPUT_PATH = sys.argv[3]

ID_COLUMNS = {
    "addressid", "conditionid",
    "diagnosisid", "dispensingid", "encounterid",
    "facilityid", "geocodeid", "immunizationid",
    "lab_facilityid", "lab_result_cm_id", "labhistoryid",
    "medadmin_providerid", "medadminid",
    "obsclin_providerid", "obsclinid", "obsgen_providerid",
    "obsgenid", "org_patid",
    "patid", "person_id", "prescribingid",
    "pro_cm_id", "proceduresid", "providerid",
    "raw_siteid", "rx_providerid", "trial_siteid",
    "trialid", "visit_id", "vitalid",
    "vx_providerid",
    "med_id",
}


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# workbook compatibility patch
_orig_convert = _openpyxl_base._convert

def _patched_convert(expected_type, value):
    if (expected_type is datetime.datetime
            and isinstance(value, datetime.date)
            and not isinstance(value, datetime.datetime)):
        return datetime.datetime.combine(value, datetime.time.min)
    return _orig_convert(expected_type, value)

_openpyxl_base._convert = _patched_convert

# load mapping
mapping = {}
with open(MAPPING_CSV, newline="", encoding="utf-8") as fh:
    reader = csv.DictReader(fh)
    for row in reader:
        col = norm(row.get("column"))
        val = norm(row.get("original_value"))
        new_id = norm(row.get("new_id"))
        if not col or not val or not new_id:
            continue
        mapping[(col.lower(), val)] = new_id

try:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(INPUT_PATH)
except Exception as e:
    print(f"  Skipping {INPUT_PATH}: could not open workbook ({e})")
    sys.exit(0)

missing = 0
replaced = 0

for ws in wb.worksheets:
    if ws.max_row < 2:
        continue

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    id_cols = []
    for i, h in enumerate(headers):
        if h is None:
            continue
        col = str(h).strip().lower()
        if col in ID_COLUMNS:
            id_cols.append((i, col))

    if not id_cols:
        continue

    for row in ws.iter_rows(min_row=2):
        for i, col in id_cols:
            v = norm(row[i].value)
            if not v:
                continue
            mapped = mapping.get((col, v))
            if mapped is None:
                missing += 1
                continue
            if str(row[i].value) != mapped:
                row[i].value = mapped
                replaced += 1

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)

print(f"  Replaced: {replaced}")
if missing > 0:
    print(f"  WARNING: {missing} values had no mapping in {MAPPING_CSV}")
print(f"  Saved: {OUTPUT_PATH}")
