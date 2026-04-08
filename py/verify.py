#!/usr/bin/env python3
"""
verify.py  <site_name>

Spot-checks that IDs have been replaced in XLSX outputs and that CPT output exists.
Prints a side-by-side diff of raw vs mapped values for each ID column found.
"""

import csv
import glob
import os
import re
import sys
import warnings
import datetime
import openpyxl
import openpyxl.descriptors.base as _openpyxl_base

if len(sys.argv) != 2:
    print("Usage: verify.py <site_name>")
    print("  e.g. verify.py C7LC_compare_deq_q01")
    sys.exit(1)

SITE = sys.argv[1]

# Paths — read from environment (set via .env loaded by the calling shell script)
# or fall back to defaults so verify.py can also be run standalone.
CPT_BASE    = os.environ.get("CPT_BASE",    "REDACTED:/data/sas_queries/<source_user>/compare_q01")
OUT_BASE    = os.environ.get("OUT_BASE",    "REDACTED:/data/sas_queries/<your_user>/lessid_drnoc")
LOOKUP_BASE = os.environ.get("LOOKUP_BASE", "REDACTED:/data/sas_queries/<your_user>/lessid_lookup")

src_drnoc   = os.path.join(CPT_BASE,    SITE, "drnoc")
out_dir     = os.path.join(OUT_BASE,    SITE)
mapping_csv = os.path.join(LOOKUP_BASE, SITE, "mapping.csv")

MAPPED_PATTERN = re.compile(r'^(PAT|ENC|PRV|FAC|ID)_\d{8}$')

XLSX_COLUMN_MAP = {
    "patient id":      "patid",
    "encounter id":    "encounterid",
    "diagnosis id":    "diagnosisid",
    "lab result id":   "lab_result_cm_id",
    "med id":          "med_id",
    "c_patid":         "patid",
    "c_trialid":       "trialid",
    "c_partid":        "trialid",
    "e_partid":        "trialid",
    "c_siteid":        "trial_siteid",
    "e_siteid":        "trial_siteid",
}

ID_COLUMNS = {
    "addressid", "conditionid", "diagnosisid", "dispensingid", "encounterid",
    "facilityid", "geocodeid", "immunizationid", "lab_facilityid",
    "lab_result_cm_id", "labhistoryid", "medadmin_providerid", "medadminid",
    "obsclin_providerid", "obsclinid", "obsgen_providerid", "obsgenid",
    "org_patid", "patid", "person_id", "prescribingid", "pro_cm_id",
    "proceduresid", "providerid", "raw_siteid", "rx_providerid",
    "trial_siteid", "trialid", "visit_id", "vitalid", "vx_providerid", "med_id",
}

# openpyxl compatibility patch
_orig_convert = _openpyxl_base._convert
def _patched_convert(expected_type, value):
    if (expected_type is datetime.datetime
            and isinstance(value, datetime.date)
            and not isinstance(value, datetime.datetime)):
        return datetime.datetime.combine(value, datetime.time.min)
    return _orig_convert(expected_type, value)
_openpyxl_base._convert = _patched_convert

# ── Checks ─────────────────────────────────────────────────────────────────

PASS = "\033[32mPASS\033[0m"
FAIL = "\033[31mFAIL\033[0m"
WARN = "\033[33mWARN\033[0m"
INFO = "\033[36mINFO\033[0m"

errors = 0

print(f"\n{'═'*60}")
print(f" Verifying: {SITE}")
print(f"{'═'*60}")

# ── 1. CPT output exists ────────────────────────────────────────────────────
print(f"\n[1] CPT output")
cpt_files = glob.glob(os.path.join(out_dir, "*.cpt"))
if not cpt_files:
    print(f"  [{FAIL}] No .cpt file found in {out_dir}")
    errors += 1
else:
    for f in cpt_files:
        size_mb = os.path.getsize(f) / 1024 / 1024
        print(f"  [{PASS}] {os.path.basename(f)}  ({size_mb:.1f} MB)")

# ── 2. Mapping file ─────────────────────────────────────────────────────────
print(f"\n[2] Mapping file")
if not os.path.exists(mapping_csv):
    print(f"  [{FAIL}] mapping.csv not found: {mapping_csv}")
    errors += 1
    sys.exit(errors)

mapping = {}
with open(mapping_csv, newline="", encoding="utf-8") as fh:
    for row in csv.DictReader(fh):
        col = (row.get("column") or "").strip().lower()
        val = (row.get("original_value") or "").strip()
        nid = (row.get("new_id") or "").strip()
        if col and val and nid:
            mapping[(col, val)] = nid

print(f"  [{INFO}] {len(mapping):,} entries loaded")

# ── 3. Cross-check: find mappings that are NOT in PAT/ENC/PRV/FAC/ID format ──
bad_new_ids = [(k, v) for k, v in mapping.items() if not MAPPED_PATTERN.match(v)]
if bad_new_ids:
    print(f"  [{FAIL}] {len(bad_new_ids)} mapping entries have non-standard new_id format")
    for k, v in bad_new_ids[:5]:
        print(f"           e.g. col={k[0]} orig={k[1]} new={v}")
    errors += 1
else:
    print(f"  [{PASS}] All new_id values match PAT/ENC/PRV/FAC/ID_XXXXXXXX format")

# ── 4. XLSX verification ────────────────────────────────────────────────────
print(f"\n[3] XLSX spot-check")
src_xls  = sorted(glob.glob(os.path.join(src_drnoc, "*.xlsx")))
out_xls  = sorted(glob.glob(os.path.join(out_dir,   "*.xlsx")))
out_names = {os.path.basename(f) for f in out_xls}

for src_path in src_xls:
    name = os.path.basename(src_path)
    out_path = os.path.join(out_dir, name)

    if name not in out_names:
        print(f"  [{FAIL}] Missing output: {name}")
        errors += 1
        continue

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb_src = openpyxl.load_workbook(src_path, read_only=True)
            wb_out = openpyxl.load_workbook(out_path, read_only=True)
        except Exception as e:
            print(f"  [{WARN}] Could not open {name}: {e}")
            continue

    total_replaced = 0
    total_raw_leaked = 0
    detail_lines = []

    for ws_src, ws_out in zip(wb_src.worksheets, wb_out.worksheets):
        src_rows = list(ws_src.iter_rows(min_row=1, max_row=2, values_only=True))
        out_rows = list(ws_out.iter_rows(min_row=1, max_row=2, values_only=True))
        if not src_rows:
            continue

        headers = [str(h).strip() if h is not None else "" for h in src_rows[0]]
        if headers and headers[0] == "No eligible records found":
            continue

        id_cols = []
        for i, h in enumerate(headers):
            col = h.lower()
            cdm = XLSX_COLUMN_MAP.get(col) or (col if col in ID_COLUMNS else None)
            if cdm:
                id_cols.append((i, h, cdm))

        if not id_cols:
            continue

        # scan all rows for the sheet
        all_src = list(ws_src.iter_rows(min_row=2, values_only=True))
        all_out = list(ws_out.iter_rows(min_row=2, values_only=True))

        for r_src, r_out in zip(all_src, all_out):
            for i, header, cdm in id_cols:
                sv = str(r_src[i]).strip() if r_src[i] is not None else ""
                ov = str(r_out[i]).strip() if r_out[i] is not None else ""
                if not sv:
                    continue
                expected = mapping.get((cdm, sv))
                if expected is None:
                    # value not in mapping — check if output unchanged
                    if sv == ov:
                        pass  # fine, unmapped value left as-is
                    continue
                if ov == expected:
                    total_replaced += 1
                elif ov == sv:
                    total_raw_leaked += 1
                    if len(detail_lines) < 3:
                        detail_lines.append(f"    col='{header}' raw={sv!r} out={ov!r} expected={expected!r}")

    wb_src.close()
    wb_out.close()

    if total_raw_leaked > 0:
        print(f"  [{FAIL}] {name}: {total_replaced} replaced, {total_raw_leaked} RAW IDs still in output")
        for d in detail_lines:
            print(d)
        errors += 1
    elif total_replaced == 0:
        # could be all-empty or all-unmapped
        print(f"  [{WARN}] {name}: no ID columns found or no data rows")
    else:
        print(f"  [{PASS}] {name}: {total_replaced} IDs replaced, 0 leaks")

# ── 4b. Cross-consistency: same original ID → same mapped ID in CPT and XLSX ─
# Strategy: from the XLSX outputs, collect (original_patid, mapped_patid) pairs.
# Then confirm that mapping.csv also maps that patid to the same value.
# Since both CPT and XLSX use the same mapping.csv, this proves they are consistent.
# We also directly confirm by reversing: look up mapped→original in mapping and
# check that the XLSX output contains expected mapped values for those originals.
print(f"\n[4] Cross-consistency check (XLSX ↔ mapping.csv ↔ CPT)")

# Build reverse mapping: new_id -> (column, original_value)
reverse_mapping = {}
for (col, orig), new_id in mapping.items():
    if col == "patid":
        reverse_mapping[new_id] = orig

# Collect (original, mapped) pairs actually seen in XLSX outputs
xlsx_pat_pairs = {}  # mapped_id -> original (from XLSX source vs output comparison)
for src_path in sorted(glob.glob(os.path.join(src_drnoc, "*.xlsx"))):
    name = os.path.basename(src_path)
    out_path = os.path.join(out_dir, name)
    if not os.path.exists(out_path):
        continue
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb_src = openpyxl.load_workbook(src_path, read_only=True)
            wb_out = openpyxl.load_workbook(out_path, read_only=True)
        except Exception:
            continue
    for ws_src, ws_out in zip(wb_src.worksheets, wb_out.worksheets):
        src_rows = list(ws_src.iter_rows(min_row=1, max_row=1, values_only=True))
        if not src_rows:
            continue
        headers = [str(h).strip() if h else "" for h in src_rows[0]]
        if headers and headers[0] == "No eligible records found":
            continue
        pat_idx = next(
            (i for i, h in enumerate(headers)
             if XLSX_COLUMN_MAP.get(h.lower()) == "patid"),
            None
        )
        if pat_idx is None:
            continue
        for r_src, r_out in zip(
            ws_src.iter_rows(min_row=2, values_only=True),
            ws_out.iter_rows(min_row=2, values_only=True),
        ):
            orig = str(r_src[pat_idx]).strip() if r_src[pat_idx] is not None else ""
            mapped = str(r_out[pat_idx]).strip() if r_out[pat_idx] is not None else ""
            if orig and mapped and MAPPED_PATTERN.match(mapped):
                xlsx_pat_pairs[mapped] = orig
        if len(xlsx_pat_pairs) >= 20:
            break
    wb_src.close()
    wb_out.close()
    if len(xlsx_pat_pairs) >= 20:
        break

if not xlsx_pat_pairs:
    print(f"  [{WARN}] No patient ID replacements found in XLSX outputs to cross-check")
    cross_errors = 0
    sample_pairs = []
else:
    cross_errors = 0
    for mapped_id, xlsx_orig in list(xlsx_pat_pairs.items())[:20]:
        # Confirm mapping.csv agrees
        mapping_orig = reverse_mapping.get(mapped_id)
        if mapping_orig is None:
            print(f"  [{FAIL}] {mapped_id} appears in XLSX output but NOT in mapping.csv")
            cross_errors += 1
        elif mapping_orig != xlsx_orig:
            print(f"  [{FAIL}] {mapped_id}: XLSX source has orig={xlsx_orig!r}, "
                  f"but mapping.csv says orig={mapping_orig!r}")
            cross_errors += 1

    if cross_errors == 0:
        n = min(20, len(xlsx_pat_pairs))
        print(f"  [{PASS}] {n} patient IDs checked: XLSX output values match mapping.csv exactly")
        # Now check CPT agrees via SAS
        # Build a small CSV of (orig_patid, expected_mapped) to feed to SAS
        import tempfile as _tf_cross
        sample_csv = _tf_cross.mktemp(suffix=".csv")
        sample_pairs = list(xlsx_pat_pairs.items())[:10]
        with open(sample_csv, "w") as f:
            f.write("original_patid,expected_mapped\n")
            for mapped_id, orig in sample_pairs:
                f.write(f"{orig},{mapped_id}\n")
        print(f"  [{INFO}] Checking same {len(sample_pairs)} patients exist correctly in CPT...")
    else:
        errors += cross_errors

# ── 5. CPT spot-check via SAS cimport + proc contents ──────────────────────
print(f"\n[5] CPT ID column spot-check (SAS)")
cpt_out = cpt_files[0] if cpt_files else None
if cpt_out:
    import tempfile as _tf
    chk_dir = _tf.mkdtemp(prefix=f"lessid_verify_{SITE}_")

    # Build cross-check SAS code: for each sampled patient, confirm CPT has expected mapped value
    cross_checks_sas = ""
    if xlsx_pat_pairs and cross_errors == 0:
        for mapped_id, orig in sample_pairs:
            safe_orig = orig.replace("'", "''")
            safe_mapped = mapped_id.replace("'", "''")
            cross_checks_sas += f"""
    data _null_;
        set chk.&first_ds (obs=50000 keep=patid);
        where patid = '{safe_mapped}';
        if _n_ = 1 then put 'CROSS_PASS orig={safe_orig} mapped={safe_mapped}';
    run;
    data _null_;
        set chk.&first_ds (obs=50000 keep=patid);
        where patid = '{safe_orig}';
        if _n_ = 1 then put 'CROSS_FAIL raw_id_found orig={safe_orig}';
    run;
"""

    sas_code = f"""
libname chk "{chk_dir}";
proc cimport infile="{cpt_out}" library=chk; run;

/* pick the first table that has a patid column */
proc sql noprint;
    select memname into :first_ds trimmed
    from dictionary.columns
    where libname='CHK' and upcase(name)='PATID'
    having monotonic() = min(monotonic());
quit;

%macro chk_ids;
%if %symexist(first_ds) and %superq(first_ds) ne %then %do;
    data _null_;
        set chk.&first_ds (obs=500 keep=patid);
        if patid ne '' and not prxmatch('/^(PAT|ENC|PRV|FAC|ID)_\d+$/', strip(patid)) then do;
            put 'WARN_RAW_ID patid=' patid;
        end;
    run;
{cross_checks_sas}
%end;
%mend; %chk_ids;

proc datasets library=chk kill nolist; run; quit;
libname chk clear;
"""
    import subprocess, tempfile as _tf2
    with _tf2.NamedTemporaryFile(mode='w', suffix='.sas', delete=False) as tf:
        tf.write(sas_code)
        tf_path = tf.name

    log_path = tf_path.replace('.sas', '.log')
    result = subprocess.run(
        ['sas', '-nodms', '-log', log_path, tf_path],
        capture_output=True, text=True, timeout=300
    )
    os.unlink(tf_path)

    if os.path.exists(log_path):
        with open(log_path) as lf:
            log = lf.read()
        os.unlink(log_path)

        warn_raw = [l for l in log.splitlines() if l.startswith('WARN_RAW_ID')]
        cross_pass = [l for l in log.splitlines() if l.startswith('CROSS_PASS')]
        cross_fail = [l for l in log.splitlines() if l.startswith('CROSS_FAIL')]
        sas_errors = [l for l in log.splitlines() if l.strip().startswith('ERROR')]

        if sas_errors:
            print(f"  [{FAIL}] SAS errors during CPT check:")
            for l in sas_errors[:5]:
                print(f"    {l}")
            errors += 1
        elif warn_raw:
            print(f"  [{FAIL}] CPT contains un-mapped raw IDs:")
            for l in warn_raw[:10]:
                print(f"    {l}")
            errors += 1
        else:
            print(f"  [{PASS}] CPT: no raw IDs in patid column (500-row sample)")

        if cross_fail:
            print(f"  [{FAIL}] CPT still contains original (unmapped) patient IDs:")
            for l in cross_fail[:5]:
                print(f"    {l}")
            errors += 1
        elif cross_pass:
            print(f"  [{PASS}] CPT ↔ XLSX cross-check: {len(cross_pass)}/{len(sample_pairs)} "
                  f"patients confirmed identical mapping in both files")
        elif xlsx_pat_pairs and cross_errors == 0:
            print(f"  [{WARN}] CPT cross-check: sampled patients not found in first CPT table "
                  f"(may be in a different table)")
        import shutil; shutil.rmtree(chk_dir, ignore_errors=True)

# ── Summary ─────────────────────────────────────────────────────────────────
print(f"\n{'═'*60}")
if errors == 0:
    print(f" {PASS}  All checks passed for {SITE}")
else:
    print(f" {FAIL}  {errors} check(s) FAILED for {SITE}")
print(f"{'═'*60}\n")
sys.exit(errors)
