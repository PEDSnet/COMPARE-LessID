#!/usr/bin/env python3
import csv
import os
import sys
import warnings
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rules import is_remap_col, mapping_col, prefix_for, XLSX_COLUMN_MAP, norm, patch_openpyxl
patch_openpyxl()

if len(sys.argv) != 6:
    print("Usage: build_site_mapping.py <site_name> <cpt_ids_csv> <xlsx_dir> <mapping_csv> <report_txt>")
    sys.exit(1)

SITE_NAME = sys.argv[1]
CPT_IDS_CSV = sys.argv[2]
XLSX_DIR = sys.argv[3]
MAPPING_CSV = sys.argv[4]
REPORT_TXT = sys.argv[5]

# Translates COMPARE study XLSX column headers (lowercased) to CDM column names
pairs = set()  # (column_lower, original_value)
cpt_pairs = 0
xlsx_pairs = 0
xlsx_files_seen = 0

# Load CPT id values extracted by SAS
if os.path.exists(CPT_IDS_CSV):
    with open(CPT_IDS_CSV, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            col = norm(row.get("column"))
            val = norm(row.get("original_value"))
            if not col or not val:
                continue
            col = col.lower()
            if not is_remap_col(col):
                continue
            key = (mapping_col(col), val)
            if key not in pairs:
                pairs.add(key)
                cpt_pairs += 1

# Load XLSX id values
for name in sorted(os.listdir(XLSX_DIR)):
    if not name.lower().endswith(".xlsx"):
        continue
    path = os.path.join(XLSX_DIR, name)
    if not os.path.isfile(path):
        continue

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path)
    except Exception:
        continue

    xlsx_files_seen += 1

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        id_idx = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            raw_col = str(h).strip().lower()
            cdm = XLSX_COLUMN_MAP.get(raw_col, raw_col)
            if is_remap_col(cdm):
                id_idx.append((i, mapping_col(cdm)))

        if not id_idx:
            continue

        for row in ws.iter_rows(min_row=2):
            for i, col in id_idx:
                val = norm(row[i].value)
                if not val:
                    continue
                key = (col, val)
                if key not in pairs:
                    pairs.add(key)
                    xlsx_pairs += 1

# Deterministic assignment order
ordered = sorted(pairs, key=lambda t: (prefix_for(t[0]), t[0], t[1]))

# Site code: everything before the first underscore (e.g. "C7LC" from "C7LC_compare_deq_q01")
site_code = SITE_NAME.split("_")[0].upper()

prefix_counters = {}
rows = []
for col, val in ordered:
    pfx = prefix_for(col)
    prefix_counters[pfx] = prefix_counters.get(pfx, 0) + 1
    new_id = f"{pfx}_{site_code}_{prefix_counters[pfx]:08d}"
    rows.append((col, val, new_id))

os.makedirs(os.path.dirname(MAPPING_CSV), exist_ok=True)
with open(MAPPING_CSV, "w", newline="", encoding="utf-8") as fh:
    writer = csv.writer(fh)
    writer.writerow(["column", "original_value", "new_id"])
    writer.writerows(rows)

os.makedirs(os.path.dirname(REPORT_TXT), exist_ok=True)
with open(REPORT_TXT, "w", encoding="utf-8") as fh:
    fh.write(f"site: {SITE_NAME}\n")
    fh.write(f"mapping_csv: {MAPPING_CSV}\n")
    fh.write(f"total_mappings: {len(rows)}\n")
    fh.write(f"from_cpt_unique_pairs: {cpt_pairs}\n")
    fh.write(f"from_xlsx_new_pairs: {xlsx_pairs}\n")
    fh.write(f"xlsx_files_seen: {xlsx_files_seen}\n")
    for pfx in sorted(prefix_counters):
        fh.write(f"prefix_{pfx}: {prefix_counters[pfx]}\n")

print(f"  Mapping rows: {len(rows)} -> {MAPPING_CSV}")
print(f"  Report: {REPORT_TXT}")
