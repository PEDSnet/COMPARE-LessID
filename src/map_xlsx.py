#!/usr/bin/env python3
import csv
import sys
import os
import warnings
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rules import is_remap_col, is_redact_col, mapping_col, XLSX_COLUMN_MAP, norm, patch_openpyxl
patch_openpyxl()

if len(sys.argv) != 4:
    print("Usage: map_xlsx.py <mapping.csv> <input.xlsx> <output.xlsx>")
    sys.exit(1)

MAPPING_CSV = sys.argv[1]
INPUT_PATH = sys.argv[2]
OUTPUT_PATH = sys.argv[3]

# load mapping
mapping = {}
with open(MAPPING_CSV, newline="", encoding="utf-8") as fh:
    reader = csv.DictReader(fh)
    for row in reader:
        col = norm(row.get("column"))
        val = norm(row.get("original_value"))
        new_id = norm(row.get("new_id"))
        if not col or not val or not new_id:
            continue
        mapping[(col.lower(), val)] = new_id

try:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(INPUT_PATH)
except Exception as e:
    print(f"  Skipping {INPUT_PATH}: could not open workbook ({e})")
    sys.exit(0)

missing = 0
replaced = 0

for ws in wb.worksheets:
    if ws.max_row < 2:
        continue

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    remap_cols = []   # (col_idx, canonical_mapping_key)
    redact_cols = []  # col_idx only
    for i, h in enumerate(headers):
        if h is None:
            continue
        raw_col = str(h).strip().lower()
        cdm = XLSX_COLUMN_MAP.get(raw_col, raw_col)
        if is_redact_col(cdm):
            redact_cols.append(i)
        elif is_remap_col(cdm):
            remap_cols.append((i, mapping_col(cdm)))

    if not remap_cols and not redact_cols:
        continue

    for row in ws.iter_rows(min_row=2):
        for i in redact_cols:
            if row[i].value is not None and row[i].value != "":
                row[i].value = ""
                replaced += 1
        for i, col in remap_cols:
            v = norm(row[i].value)
            if not v:
                continue
            mapped = mapping.get((col, v))
            if mapped is None:
                missing += 1
                continue
            if str(row[i].value) != mapped:
                row[i].value = mapped
                replaced += 1

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)

print(f"  Replaced: {replaced}")
if missing > 0:
    print(f"  WARNING: {missing} values had no mapping in {MAPPING_CSV}")
print(f"  Saved: {OUTPUT_PATH}")
