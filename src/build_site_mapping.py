#!/usr/bin/env python3
"""
build_site_mapping.py  <site_name> <cpt_ids_csv> <xlsx_dir> <mapping_csv> <report_txt> <site_meta_csv> <query_version>

Builds (or cumulatively extends) the site-level mapping.csv.

site_code is read from site_meta.csv (DATAMARTID from HARVEST) — authoritative.
query_version is passed explicitly from the pipeline config (e.g. "1").
Folder-name parsing is used only as a last-resort fallback if site_meta.csv is missing.

The mapping is cumulative across queries:
  - Existing rows are never changed (surrogate IDs are immutable).
  - New IDs from the current query are appended and tagged with the current query.
  - Per-prefix counters continue from the highest number already in the file.

mapping.csv columns: column, original_value, new_id, query
"""
import csv
import os
import re
import sys
import warnings
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rules import is_remap_col, mapping_col, prefix_for, XLSX_COLUMN_MAP, norm, patch_openpyxl
patch_openpyxl()

if len(sys.argv) != 8:
    print("Usage: build_site_mapping.py <site_name> <cpt_ids_csv> <xlsx_dir> <mapping_csv> <report_txt> <site_meta_csv> <query_version>")
    sys.exit(1)

SITE_NAME     = sys.argv[1]
CPT_IDS_CSV   = sys.argv[2]
XLSX_DIR      = sys.argv[3]
MAPPING_CSV   = sys.argv[4]
REPORT_TXT    = sys.argv[5]
SITE_META_CSV = sys.argv[6]
QUERY_VERSION = sys.argv[7]  # integer as string, e.g. "1"

# ── Determine site_code and query ─────────────────────────────────────────────
# site_code: read from site_meta.csv (DATAMARTID from HARVEST) — authoritative
# query:     from query_version arg (passed from config) — no folder-name parsing
query = QUERY_VERSION
site_code = None
if os.path.exists(SITE_META_CSV):
    with open(SITE_META_CSV, newline='', encoding='utf-8') as _fh:
        _row = next(csv.DictReader(_fh), None)
        if _row:
            site_code = (_row.get('datamartid') or '').strip().upper()
if not site_code:
    print(f"WARNING: Could not read datamartid from {SITE_META_CSV!r}. "
          "Falling back to folder name.", file=sys.stderr)
    _m = re.match(r'^(?P<site>.+?)_compare.*?_(q\d+)$', SITE_NAME, re.IGNORECASE)
    if not _m:
        print(f"ERROR: Cannot parse site code from folder name: {SITE_NAME!r}", file=sys.stderr)
        sys.exit(1)
    site_code = _m.group('site').upper()

# ── Load existing mapping (cumulative) ───────────────────────────────────────
# existing: dict[(column, original_value)] -> (new_id, query_first_seen)
existing = {}
prefix_max = {}   # prefix -> highest sequence number already assigned

if os.path.exists(MAPPING_CSV):
    with open(MAPPING_CSV, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            col = (row.get("column") or "").strip().lower()
            val = (row.get("original_value") or "").strip()
            nid = (row.get("new_id") or "").strip()
            q   = (row.get("query") or "").strip()
            if not (col and val and nid):
                continue
            existing[(col, val)] = (nid, q)
            # Track highest sequence number per prefix so new IDs continue from there
            _seq_m = re.match(r'^([A-Z]+)_[A-Z0-9]+_(\d+)$', nid)
            if _seq_m:
                pfx = _seq_m.group(1)
                seq = int(_seq_m.group(2))
                if seq > prefix_max.get(pfx, 0):
                    prefix_max[pfx] = seq

# ── Collect (column, value) pairs from current query ─────────────────────────
new_pairs = set()
cpt_pairs = 0
xlsx_pairs = 0
xlsx_files_seen = 0

if os.path.exists(CPT_IDS_CSV):
    with open(CPT_IDS_CSV, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            col = norm(row.get("column"))
            val = norm(row.get("original_value"))
            if not col or not val:
                continue
            col = col.lower()
            if not is_remap_col(col):
                continue
            key = (mapping_col(col), val)
            if key not in existing and key not in new_pairs:
                new_pairs.add(key)
                cpt_pairs += 1

for name in sorted(os.listdir(XLSX_DIR)):
    if not name.lower().endswith(".xlsx"):
        continue
    path = os.path.join(XLSX_DIR, name)
    if not os.path.isfile(path):
        continue
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path)
    except Exception:
        continue
    xlsx_files_seen += 1
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        id_idx = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            raw_col = str(h).strip().lower()
            cdm = XLSX_COLUMN_MAP.get(raw_col, raw_col)
            if is_remap_col(cdm):
                id_idx.append((i, mapping_col(cdm)))
        if not id_idx:
            continue
        for row in ws.iter_rows(min_row=2):
            for i, col in id_idx:
                val = norm(row[i].value)
                if not val:
                    continue
                key = (col, val)
                if key not in existing and key not in new_pairs:
                    new_pairs.add(key)
                    xlsx_pairs += 1

# ── Assign surrogate IDs to new pairs ────────────────────────────────────────
# Deterministic order: by prefix, column, value
ordered_new = sorted(new_pairs, key=lambda t: (prefix_for(t[0]), t[0], t[1]))

prefix_counters = dict(prefix_max)   # start from existing maxima
new_rows = []
for col, val in ordered_new:
    pfx = prefix_for(col)
    prefix_counters[pfx] = prefix_counters.get(pfx, 0) + 1
    new_id = f"{pfx}_{site_code}_{prefix_counters[pfx]:08d}"
    new_rows.append((col, val, new_id, query))

# ── Write merged mapping ──────────────────────────────────────────────────────
os.makedirs(os.path.dirname(MAPPING_CSV), exist_ok=True)
with open(MAPPING_CSV, "w", newline="", encoding="utf-8") as fh:
    writer = csv.writer(fh)
    writer.writerow(["column", "original_value", "new_id", "query"])
    # Existing rows first (preserve order), then new
    for (col, val), (nid, q) in existing.items():
        writer.writerow([col, val, nid, q])
    writer.writerows(new_rows)

total = len(existing) + len(new_rows)
os.makedirs(os.path.dirname(REPORT_TXT), exist_ok=True)
with open(REPORT_TXT, "w", encoding="utf-8") as fh:
    fh.write(f"site: {site_code}\n")
    fh.write(f"query: {query}\n")
    fh.write(f"mapping_csv: {MAPPING_CSV}\n")
    fh.write(f"total_mappings: {total}\n")
    fh.write(f"existing_mappings: {len(existing)}\n")
    fh.write(f"new_this_query: {len(new_rows)}\n")
    fh.write(f"from_cpt_new_pairs: {cpt_pairs}\n")
    fh.write(f"from_xlsx_new_pairs: {xlsx_pairs}\n")
    fh.write(f"xlsx_files_seen: {xlsx_files_seen}\n")
    for pfx in sorted(prefix_counters):
        fh.write(f"prefix_{pfx}: {prefix_counters[pfx]}\n")

print(f"  Mapping rows: {total} ({len(new_rows)} new this query) -> {MAPPING_CSV}")
print(f"  Report: {REPORT_TXT}")
